import openpyxl


class RetrieveExcelData:
    sheetName= None
    @staticmethod
    def getTestData(test_case_name, sheetName):
        Dict = {}
        book = openpyxl.load_workbook("C://Users//ALAGU SUNDARI//PycharmProjects//IQ-CRM_Admin//TestData//TestData.xlsx")
       #sheet = book.active
        sheet = book[sheetName]
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(1, sheet.max_column + 1):  # to get columns
                    # Dict["firstName"]="aravindhan"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]

